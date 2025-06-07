# founder_equity_calculator.py

import streamlit as st
import pandas as pd
import openpyxl
import numpy as np
from typing import Dict, Tuple, List
import plotly.graph_objs as go

# ========== CONFIGURATION ==========
EXCEL_PATH = '20250602_Capitalisation-table_Model_InnovaIVF.xlsx'
SHEET_NAME = 'Cap Table Modelling_InnovaIVF'

# Blue theme colours in file (adjust as needed)
THEME_BLUES = {"FF9DC3E6", "FFC6E0B4"}  # e.g. light blue fills
# Excel "theme" fgColor is detected by .type == 'theme', some files use rgb as above

# ========== HELPER FUNCTIONS ==========

def load_workbook_dual(path: str, sheet: str):
    """Load workbook twice: once for formulas, once for values."""
    wb_formula = openpyxl.load_workbook(path, data_only=False)
    wb_values = openpyxl.load_workbook(path, data_only=True)
    ws_formula = wb_formula[sheet]
    ws_values = wb_values[sheet]
    return ws_formula, ws_values

def get_editable_cells(ws_formula, ws_values) -> Dict[str, dict]:
    """
    Find all blue solid-fill cells (editable) and their labels, addresses and current values.
    Returns dict keyed by Excel address, with keys: 'row', 'col', 'address', 'label', 'value'.
    """
    editable = {}
    max_col = ws_formula.max_column
    for row in ws_formula.iter_rows():
        for cell in row:
            fill = cell.fill
            # Detect solid blue theme fill
            if fill and fill.patternType == 'solid':
                fc = fill.fgColor
                is_theme = (getattr(fc, 'type', None) == 'theme')
                is_blue = (getattr(fc, 'rgb', None) in THEME_BLUES)
                if is_theme or is_blue:
                    address = cell.coordinate
                    col_idx = cell.column
                    row_idx = cell.row
                    # Try to get label: check cell to left, else column header
                    if col_idx > 1:
                        label = ws_formula.cell(row=row_idx, column=col_idx-1).value
                    else:
                        # Use header from row 1
                        label = ws_formula.cell(row=1, column=col_idx).value
                    value = ws_values[address].value
                    editable[address] = {
                        'row': row_idx, 'col': col_idx,
                        'address': address, 'label': label or address, 'value': value
                    }
    return editable

def parse_inputs(editable_cells: Dict[str, dict]) -> Dict[str, float]:
    """
    Parse and cast user-edited values from widgets into dictionary, keyed by cell address.
    """
    parsed = {}
    for addr, cell in editable_cells.items():
        v = st.session_state.get(f"cell_{addr}", cell['value'])
        try:
            parsed[addr] = float(v)
        except (TypeError, ValueError):
            parsed[addr] = v  # Allow string for round names etc
    return parsed

def extract_model_inputs(ws_formula, ws_values, editable_addrs: List[str]) -> Tuple[List[str], Dict[str, dict]]:
    """
    Parse entire sheet for model logic: rounds, stakeholders, cap table shape, etc.
    """
    # Find rounds as columns (usually header row)
    header_row = 1
    rounds = []
    for col in range(2, ws_formula.max_column+1):
        name = ws_formula.cell(row=header_row, column=col).value
        if name:
            rounds.append(name)
    # Find stakeholders as rows (first column)
    stakeholders = []
    for row in range(2, ws_formula.max_row+1):
        label = ws_formula.cell(row=row, column=1).value
        if label:
            stakeholders.append(label)
    return rounds, stakeholders

def recalculate_cap_table(inputs: Dict[str, float], rounds: List[str], stakeholders: List[str]) -> Tuple[pd.DataFrame, pd.DataFrame, dict]:
    """
    Core business logic for cap table calculation. Simulates share issuances, valuations, dilution.
    Returns:
      - ownership table (rows: stakeholders, columns: rounds)
      - shares table (absolute number of shares)
      - summary dict (headline results)
    """
    # Business rules must be inferred from sheet structure and provided labels
    # For demo, assume:
    #  - Inputs contain: round names, pre-money valuation, investment amount, ESOP size, founders' shares
    #  - Stakeholders: Founders, ESOP, Investors, Unallocated pool, etc.

    # Example input mapping:
    # Let's map: Pre-money (by round), Investment (by round), ESOP %, Initial Shares, etc.
    # These would have been given "labels" (e.g. "Pre-Money Seed"), parse as such.
    # For clarity, below is a typical round structure:
    # rounds = ['Foundation', 'Pre-Seed', 'Seed', 'Series A', 'Series B', 'Exit']
    # Each round: record pre-money, new $ in, ESOP, share price, new shares issued
    n_rounds = len(rounds)
    n_founders = [s for s in stakeholders if 'Founder' in s]
    cap_table = pd.DataFrame(0.0, index=stakeholders, columns=rounds)
    shares = pd.DataFrame(0.0, index=stakeholders, columns=rounds)
    all_round_data = []
    
    # Business logic placeholder (replace this with formulae adapted from your sheet)
    # --- EXAMPLE LOGIC (adapt as needed based on sheet):
    pre_money = []
    invest = []
    esop = []
    founder_shares_init = []
    for r in rounds:
        for addr, cell in inputs.items():
            if r.lower() in str(addr).lower() and 'pre' in str(addr).lower():
                pre_money.append(float(inputs[addr]))
            if r.lower() in str(addr).lower() and 'invest' in str(addr).lower():
                invest.append(float(inputs[addr]))
            if r.lower() in str(addr).lower() and 'esop' in str(addr).lower():
                esop.append(float(inputs[addr]))
        # Founders' initial shares - for Foundation
        for addr, cell in inputs.items():
            if 'founder' in str(addr).lower() and (r == rounds[0]):
                founder_shares_init.append(float(inputs[addr]))
    # Fallback defaults if not enough values found
    while len(pre_money) < n_rounds:
        pre_money.append(1e6)
    while len(invest) < n_rounds:
        invest.append(0)
    while len(esop) < n_rounds:
        esop.append(0.10)

    # Set up shares
    total_shares = [sum(founder_shares_init) if founder_shares_init else 1000000]
    share_price = [pre_money[0] / total_shares[0]]
    # Fill initial founder stakes
    for i, s in enumerate(stakeholders):
        if 'Founder' in s:
            shares.loc[s, rounds[0]] = founder_shares_init[i] if i < len(founder_shares_init) else total_shares[0] / len(n_founders)
    # Loop through rounds
    for i in range(1, n_rounds):
        # Compute post-money, new shares
        total_prev = shares[rounds[i-1]].sum()
        post_money = pre_money[i] + invest[i]
        share_price.append(pre_money[i] / total_prev if total_prev else 1.0)
        new_shares = invest[i] / share_price[i] if share_price[i] else 0
        # ESOP allocation (simple % of post)
        esop_shares = esop[i] * (total_prev + new_shares) / (1 - esop[i]) if esop[i] else 0
        # Allocate shares
        for s in stakeholders:
            # Founders get diluted, unless protected
            if 'Founder' in s:
                prev = shares.loc[s, rounds[i-1]]
                shares.loc[s, rounds[i]] = prev
            elif 'ESOP' in s:
                shares.loc[s, rounds[i]] = esop_shares
            elif 'Investor' in s:
                shares.loc[s, rounds[i]] = new_shares / (len([k for k in stakeholders if 'Investor' in k]) or 1)
        # Sum up for next round
        total_shares.append(shares[rounds[i]].sum())
    # Ownership %
    for r in rounds:
        total = shares[r].sum()
        if total == 0:
            continue
        for s in stakeholders:
            cap_table.loc[s, r] = shares.loc[s, r] / total * 100

    # Example headline summary
    founder_exit = {s: cap_table.loc[s, rounds[-1]] for s in stakeholders if 'Founder' in s}
    investor_irr = 0.35  # Placeholder, compute from $ invested and exit value
    summary = {
        "Founders' ownership at exit": founder_exit,
        "Investor IRR": f"{investor_irr*100:.1f}%",
        "Money-on-money": f"{(pre_money[-1]/(sum(invest) or 1)):.2f}x"
    }
    return cap_table, shares, summary

def build_ui(editable_cells, rounds, stakeholders, cap_table, shares, summary):
    st.title("Founder Equity Calculator")
    st.write("This tool calculates fully-diluted ownership through multiple investment rounds, based on your Excel model.")

    st.header("Model Inputs")
    with st.form("inputs_form"):
        for addr, meta in editable_cells.items():
            label = meta['label']
            value = meta['value']
            if isinstance(value, (int, float)):
                st.number_input(label, value=value, key=f"cell_{addr}")
            else:
                st.text_input(label, value=str(value), key=f"cell_{addr}")
        submitted = st.form_submit_button("Update Cap Table")

    st.header("Cap Table by Stakeholder and Round")
    st.dataframe(cap_table.style.format("{:.2f}"))

    st.header("Founders' Ownership Over Time")
    founder_rows = [s for s in stakeholders if 'Founder' in s]
    fig = go.Figure()
    for s in founder_rows:
        fig.add_trace(go.Scatter(
            x=cap_table.columns, y=cap_table.loc[s], mode='lines+markers', name=s
        ))
    fig.update_layout(
        xaxis_title="Round",
        yaxis_title="Ownership (%)",
        legend_title="Founder",
        template="plotly_white"
    )
    st.plotly_chart(fig)

    st.header("Headline Results")
    for k, v in summary.items():
        st.markdown(f"**{k}:** {v}")

# ========== MAIN APP FLOW ==========

def main():
    # Load and parse workbook
    ws_formula, ws_values = load_workbook_dual(EXCEL_PATH, SHEET_NAME)
    editable_cells = get_editable_cells(ws_formula, ws_values)
    rounds, stakeholders = extract_model_inputs(ws_formula, ws_values, list(editable_cells.keys()))
    # Collect user inputs and update
    if "cells_loaded" not in st.session_state:
        for addr, meta in editable_cells.items():
            st.session_state[f"cell_{addr}"] = meta['value']
        st.session_state["cells_loaded"] = True
    user_inputs = parse_inputs(editable_cells)
    # Recalculate cap table
    cap_table, shares, summary = recalculate_cap_table(user_inputs, rounds, stakeholders)
    # Build UI
    build_ui(editable_cells, rounds, stakeholders, cap_table, shares, summary)

if __name__ == "__main__":
    main()
